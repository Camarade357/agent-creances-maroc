"""
Moteur d'analyse Balance Agee - traitement 100% en memoire (zero Supabase).
Detecte automatiquement les formats BEST MILK (FORMAT_A) et SWF/FANDY (FORMAT_B).
"""
import datetime
import openpyxl
import pandas as pd
from io import BytesIO

_AMOUNT_COLS = ['non_echu', 'b_0_30', 'b_30_60', 'b_60_90', 'b_90_120', 'b_plus120', 'avances', 'solde']


def _abs_amounts(df):
    """Retourne une copie du DataFrame avec les colonnes monetaires en abs()."""
    out = df.copy()
    for c in _AMOUNT_COLS:
        if c in out.columns:
            out[c] = out[c].abs()
    return out


def to_float(val):
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def detect_format(ws):
    """
    FORMAT_A (BEST MILK) : headers contiennent 'RaisonSocial' / 'Tier' / 'Solde'
    FORMAT_B (SWF/FANDY) : headers contiennent 'Non echu' / 'Plus de 90'
    FORMAT_C (Bce aux)   : headers contiennent SFD / SFC / mouvements (balance auxiliaire,
                           pas de buckets d'age - non exploitable pour aging)
    Les headers sont parfois en L2 (export comptable avec ligne titre vide en L1).
    On lit L1 et L2 et on normalise pour matcher.
    """
    parts = []
    for row in ws.iter_rows(min_row=1, max_row=2, values_only=True):
        parts.extend(str(h).strip() if h else '' for h in row)
    headers_norm = ' '.join(parts).lower()
    headers_norm = (headers_norm
                    .replace('é', 'e').replace('è', 'e').replace('ê', 'e')
                    .replace('à', 'a').replace('°', ''))

    if 'raisonsocial' in headers_norm or ('tier' in headers_norm and 'solde' in headers_norm):
        return 'FORMAT_A'
    if 'non echu' in headers_norm or 'plus de 90' in headers_norm:
        return 'FORMAT_B'
    if any(k in headers_norm for k in [
        'sfd', 'sfc', 'solde final', 'mouvement',
        "solde d'ouverture", 'solde ouverture',
    ]) or ('debit' in headers_norm and 'credit' in headers_norm):
        return 'FORMAT_C'
    return 'FORMAT_UNKNOWN'


def find_data_sheet(wb):
    """
    Parcourt les onglets et retourne le premier ayant un format reconnu.
    Si aucun reconnu, retourne le premier onglet ayant >= 5 headers non vides.
    """
    fallback = (None, None)
    for name in wb.sheetnames:
        ws = wb[name]
        fmt = detect_format(ws)
        if fmt in ('FORMAT_A', 'FORMAT_B', 'FORMAT_C'):
            return ws, name
        headers_count = 0
        for row in ws.iter_rows(min_row=1, max_row=2, values_only=True):
            headers_count = max(headers_count, sum(1 for h in row if h is not None))
        if fallback[0] is None and headers_count >= 5:
            fallback = (ws, name)
    return fallback


def parse_excel(file_bytes):
    """
    Parse un fichier Excel Balance Agee.
    Retourne (df, metadata) ou raise ValueError si format non reconnu.
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws, sheet_name = find_data_sheet(wb)

    if ws is None:
        wb.close()
        raise ValueError("Aucun onglet de donnees reconnu dans ce fichier.")

    fmt = detect_format(ws)
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    records = []
    skipped = 0

    if fmt == 'FORMAT_A':
        # BEST MILK : Societe/Categorie/Tier/RaisonSocial/Compte/>120/90-120/60-90/30-60/0-30/Avances/Solde
        # Valeurs SIGNEES conservees (createurs/crediteurs s'agregent correctement).
        for row in rows:
            if not row or len(row) < 12 or not row[3]:
                skipped += 1
                continue
            records.append({
                'client_id':    str(row[2]).strip() if row[2] else '',
                'client_name':  str(row[3]).strip(),
                'categorie':    str(row[1]).strip() if row[1] else '',
                'non_echu':     0.0,
                'b_0_30':       to_float(row[9]),
                'b_30_60':      to_float(row[8]),
                'b_60_90':      to_float(row[7]),
                'b_90_120':     to_float(row[6]),
                'b_plus120':    to_float(row[5]),
                'avances':      to_float(row[10]),
                'solde':        to_float(row[11]),
            })

    elif fmt == 'FORMAT_B':
        # SWF : N/Nom/Non echu/0-30/31-60/61-90/Plus de 90
        # Valeurs SIGNEES conservees ; solde recalcule = somme signee des buckets.
        for row in rows:
            if not row or len(row) < 7 or not row[0] or not row[1]:
                skipped += 1
                continue
            non_echu = to_float(row[2])
            b_0_30   = to_float(row[3])
            b_31_60  = to_float(row[4])
            b_61_90  = to_float(row[5])
            b_plus90 = to_float(row[6])
            solde    = non_echu + b_0_30 + b_31_60 + b_61_90 + b_plus90
            records.append({
                'client_id':    str(row[0]).strip(),
                'client_name':  str(row[1]).strip(),
                'categorie':    '',
                'non_echu':     non_echu,
                'b_0_30':       b_0_30,
                'b_30_60':      b_31_60,
                'b_60_90':      b_61_90,
                'b_90_120':     0.0,
                'b_plus120':    b_plus90,
                'avances':      0.0,
                'solde':        solde,
            })
    elif fmt == 'FORMAT_C':
        raise ValueError('FORMAT_C_DETECTED')
    else:
        raise ValueError('FORMAT_UNKNOWN')

    df = pd.DataFrame(records)

    metadata = {
        'format':      fmt,
        'sheet_name':  sheet_name,
        'source':      sheet_name,
        'nb_clients':  len(records),
        'nb_skipped':  skipped,
        'total_due':   float(df['solde'].abs().sum()) if len(df) else 0.0,
        'plus120':     float(df['b_plus120'].abs().sum()) if len(df) else 0.0,
    }
    metadata['pct_plus120'] = (
        metadata['plus120'] / metadata['total_due'] * 100
        if metadata['total_due'] > 0 else 0.0
    )

    return df, metadata


def compute_aging(df):
    """
    KPIs aging - agregation en VALEUR ABSOLUE par client.
    - total      = somme des |solde| par client (volume expose total)
    - buckets[k] = somme des |valeur| par client dans chaque tranche d'age
    - pct_crit   = bucket >120j / somme totale des buckets (toujours <= 100%)
    Cette convention est robuste aux ecritures inverses (montants signes mixtes).
    """
    bucket_keys = ['Non echu', '0-30 j', '30-60 j', '60-90 j', '90-120 j', '> 120 j']

    if not len(df):
        return {
            'total':         0.0,
            'buckets':       {k: 0.0 for k in bucket_keys},
            'total_buckets': 0.0,
            'top10':         pd.DataFrame(columns=['Code','Client','>120j (MAD)','Total du (MAD)']),
            'nb_critiques':  0,
            'pct_critique':  0.0,
        }

    total = float(df['solde'].abs().sum())

    buckets = {
        'Non echu':  float(df['non_echu'].abs().sum()),
        '0-30 j':    float(df['b_0_30'].abs().sum()),
        '30-60 j':   float(df['b_30_60'].abs().sum()),
        '60-90 j':   float(df['b_60_90'].abs().sum()),
        '90-120 j':  float(df['b_90_120'].abs().sum()),
        '> 120 j':   float(df['b_plus120'].abs().sum()),
    }
    total_buckets = sum(buckets.values())

    # Top 10 : clients avec b_plus120 != 0, tri par |b_plus120| DESC.
    # Affichage en valeur absolue.
    has_plus120 = df[df['b_plus120'] != 0].copy()
    if len(has_plus120):
        has_plus120['abs_plus120'] = has_plus120['b_plus120'].abs()
        top10 = (
            has_plus120
            .sort_values('abs_plus120', ascending=False)
            .head(10)[['client_id', 'client_name', 'b_plus120', 'solde']]
            .copy()
        )
        top10['b_plus120'] = top10['b_plus120'].abs()
        top10['solde']     = top10['solde'].abs()
        top10 = top10.rename(columns={
            'client_id':   'Code',
            'client_name': 'Client',
            'b_plus120':   '>120j (MAD)',
            'solde':       'Total du (MAD)',
        }).reset_index(drop=True)
        top10.index += 1
    else:
        top10 = pd.DataFrame(columns=['Code', 'Client', '>120j (MAD)', 'Total du (MAD)'])

    nb_critiques = int((df['b_plus120'] != 0).sum())

    return {
        'total':         total,
        'buckets':       buckets,
        'total_buckets': total_buckets,
        'top10':         top10,
        'nb_critiques':  nb_critiques,
        'pct_critique':  (buckets['> 120 j'] / total_buckets * 100) if total_buckets > 0 else 0.0,
    }


def compute_segmentation(df):
    """
    Segmentation ABC Pareto 80/20 x comportement paiement.

    TAILLE (grand/petit) :
      - Trier les clients par solde decroissant
      - Cumuler les soldes jusqu'a 80% du total
      - Les clients dans ce top = GRAND COMPTE
      - Les autres = PETIT COMPTE

    COMPORTEMENT PAIEMENT :
      - Bon payeur     : >=70% du solde dans non_echu + b_0_30 + b_30_60
      - Mauvais payeur : >=50% du solde dans b_plus120
      - Neutre         : entre les deux

    QUADRANTS :
      A-BON    : Grand + Bon    -> Fideliser
      A-RISQUE : Grand + Mauvais -> Priorite #1 / Appel DG
      B-BON    : Petit + Bon    -> Relance automatique
      B-RISQUE : Petit + Mauvais -> Arbitrage cout/benefice
      NEUTRE   : tout profil neutre

    Travaille sur abs(montants).
    Retourne (segments_dict, df_with_segment, pareto_info).
    """
    if df.empty:
        return {}, df, {'nb_grands': 0, 'pct_clients': 0, 'seuil_mad': 0.0, 'pct_montant': 0.0}

    df = _abs_amounts(df)
    total_global = float(df['solde'].sum())
    if total_global == 0:
        return {}, df, {'nb_grands': 0, 'pct_clients': 0, 'seuil_mad': 0.0, 'pct_montant': 0.0}

    # ── TAILLE : Pareto 80/20 ──
    df_sorted = df.sort_values('solde', ascending=False).copy()
    df_sorted['solde_cumul'] = df_sorted['solde'].cumsum()
    df_sorted['pct_cumul']   = df_sorted['solde_cumul'] / total_global
    grands_idx = df_sorted[df_sorted['pct_cumul'] < 0.80].index.tolist()
    # Inclure le premier client qui depasse 80% (couvre exactement le seuil)
    seuil_pareto_idx = df_sorted[df_sorted['pct_cumul'] >= 0.80].index
    if len(seuil_pareto_idx) > 0:
        grands_idx.append(seuil_pareto_idx[0])

    df = df.copy()
    df['size_class'] = df.index.map(
        lambda i: 'grand' if i in grands_idx else 'petit'
    )

    nb_grands = len(grands_idx)
    grands_subset = df[df['size_class'] == 'grand']
    seuil_grand = float(grands_subset['solde'].min()) if nb_grands > 0 else 0.0
    pct_grands  = round(nb_grands / len(df) * 100, 1) if len(df) > 0 else 0.0

    # ── COMPORTEMENT PAIEMENT ──
    def classify_payment(row):
        if row['solde'] == 0:
            return 'neutre'
        pct_recent   = (row.get('non_echu', 0) + row['b_0_30'] + row['b_30_60']) / row['solde']
        pct_critique = row['b_plus120'] / row['solde']
        if pct_recent >= 0.70:
            return 'bon'
        if pct_critique >= 0.50:
            return 'mauvais'
        return 'neutre'

    df['payment_class'] = df.apply(classify_payment, axis=1)

    # ── SEGMENT FINAL ──
    def segment(row):
        s, p = row['size_class'], row['payment_class']
        if s == 'grand' and p == 'bon':     return 'A-BON'
        if s == 'grand' and p == 'mauvais': return 'A-RISQUE'
        if s == 'petit' and p == 'bon':     return 'B-BON'
        if s == 'petit' and p == 'mauvais': return 'B-RISQUE'
        return 'NEUTRE'

    df['segment'] = df.apply(segment, axis=1)

    segments = {}
    for seg, label, couleur, priorite, action in [
        ('A-RISQUE', '🔴 Grands comptes — Mauvais payeurs',
         '#dc2626', 1,
         'PRIORITÉ #1 — Appel direction + mise en demeure immédiate'),
        ('B-RISQUE', '🟠 Petits comptes — Mauvais payeurs',
         '#ea580c', 2,
         'Arbitrage — coût de recouvrement vs montant dû'),
        ('A-BON',    '🟢 Grands comptes — Bons payeurs',
         '#16a34a', 3,
         'Fidéliser — conditions préférentielles, délais négociés'),
        ('B-BON',    '🟡 Petits comptes — Bons payeurs',
         '#ca8a04', 4,
         'Relance automatique standard'),
        ('NEUTRE',   '⚪ Profil neutre — À surveiller',
         '#6b7280', 5,
         'Réévaluer au prochain arrêté comptable'),
    ]:
        subset = df[df['segment'] == seg].copy()
        segments[seg] = {
            'label':     label,
            'couleur':   couleur,
            'priorite':  priorite,
            'action':    action,
            'nb':        len(subset),
            'total':     float(subset['solde'].sum()),
            'critique':  float(subset['b_plus120'].sum()),
            'pct_total': round(subset['solde'].sum() / total_global * 100, 1)
                          if total_global > 0 else 0.0,
            'clients':   subset[['client_id', 'client_name',
                                 'solde', 'b_plus120', 'b_0_30']]
                          .sort_values('solde', ascending=False)
                          .reset_index(drop=True),
        }

    pareto_info = {
        'nb_grands':   nb_grands,
        'pct_clients': pct_grands,
        'seuil_mad':   seuil_grand,
        'pct_montant': round(
            float(grands_subset['solde'].sum()) / total_global * 100, 1
        ) if total_global > 0 else 0.0,
    }

    return segments, df, pareto_info


def compute_dso(df, ca_annuel=None):
    """
    DSO approche  : Sum(montant_bucket * jours_milieu) / total
    DSO reel      : (total_creances / ca_annuel) * 365 si ca_annuel fourni
    DSO par client: meme formule par ligne.
    Travaille sur abs(montants) pour eviter les DSO negatifs.
    """
    if df.empty:
        return {
            'dso_approche': 0.0,
            'dso_reel':     None,
            'top_dso':      pd.DataFrame(columns=['Code', 'Client', 'DSO (jours)', 'Total du (MAD)', '>120j (MAD)']),
            'df_with_dso':  df,
        }

    df = _abs_amounts(df)
    total = float(df['solde'].sum())
    if total == 0:
        return {
            'dso_approche': 0.0,
            'dso_reel':     None,
            'top_dso':      pd.DataFrame(columns=['Code', 'Client', 'DSO (jours)', 'Total du (MAD)', '>120j (MAD)']),
            'df_with_dso':  df,
        }

    MILIEU = {
        'non_echu':  0,
        'b_0_30':    15,
        'b_30_60':   45,
        'b_60_90':   75,
        'b_90_120':  105,
        'b_plus120': 150,
    }

    dso_approche = sum(
        df[col].sum() * jours
        for col, jours in MILIEU.items()
        if col in df.columns
    ) / total

    dso_reel = (total / ca_annuel * 365) if ca_annuel and ca_annuel > 0 else None

    def per_client(r):
        if r['solde'] <= 0:
            return 0.0
        return sum(r.get(col, 0) * j for col, j in MILIEU.items()) / r['solde']

    df['dso_client'] = df.apply(per_client, axis=1).round(1)

    top_dso = (
        df[df['solde'] > 0]
        .sort_values('dso_client', ascending=False)
        .head(10)[['client_id', 'client_name', 'dso_client', 'solde', 'b_plus120']]
        .rename(columns={
            'client_id':   'Code',
            'client_name': 'Client',
            'dso_client':  'DSO (jours)',
            'solde':       'Total du (MAD)',
            'b_plus120':   '>120j (MAD)',
        })
        .reset_index(drop=True)
    )
    top_dso.index += 1

    return {
        'dso_approche': round(dso_approche, 1),
        'dso_reel':     round(dso_reel, 1) if dso_reel else None,
        'top_dso':      top_dso,
        'df_with_dso':  df,
    }


def generate_plan_recouvrement(segments, dso_data, date_semaine=None):
    """
    Genere un plan de recouvrement hebdomadaire structure (max 10 clients par segment).
    Priorite : B > D > A > C.
    Retourne un DataFrame exportable.
    """
    if date_semaine is None:
        date_semaine = datetime.date.today().strftime('%d/%m/%Y')

    rows = []
    priorite = 1

    action_map = {
        'A-RISQUE': 'Appel direction + email mise en demeure',
        'B-RISQUE': 'Évaluer coût recouvrement vs montant',
        'A-BON':    'Email fidélisation + conditions préférentielles',
        'B-BON':    'Relance email standard',
        'NEUTRE':   'Surveiller — prochain arrêté',
    }

    df_dso = dso_data.get('df_with_dso', pd.DataFrame())

    for seg in ['A-RISQUE', 'B-RISQUE', 'A-BON', 'B-BON', 'NEUTRE']:
        if seg not in segments:
            continue
        data = segments[seg]
        clients = data['clients'].head(10)
        for _, client in clients.iterrows():
            if client['solde'] <= 0:
                continue
            dso_val = '-'
            if not df_dso.empty and 'dso_client' in df_dso.columns:
                match = df_dso[df_dso['client_id'] == client.get('client_id', '')]
                if len(match):
                    dso_val = round(float(match['dso_client'].values[0]), 0)

            rows.append({
                'Priorité':           priorite,
                'Semaine':            date_semaine,
                'Segment':            f"{seg} · {data['label']}",
                'Code client':        client.get('client_id', ''),
                'Nom client':         client.get('client_name', ''),
                'Solde total (MAD)':  round(float(client['solde']), 2),
                '>120j (MAD)':        round(float(client.get('b_plus120', 0)), 2),
                'DSO client (j)':     dso_val,
                'Action recommandée': action_map.get(seg, 'Surveiller'),
                'Statut':             'À traiter',
            })
            priorite += 1

    return pd.DataFrame(rows)
